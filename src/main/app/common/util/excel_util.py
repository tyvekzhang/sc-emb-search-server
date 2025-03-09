import io
from datetime import datetime
from typing import List, Type
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from loguru import logger
from pydantic import BaseModel
from starlette.responses import StreamingResponse


async def export_excel(schema: Type[BaseModel], file_name: str, data_list: List[BaseModel] = []) -> StreamingResponse:
    """
    Export a template or data as an Excel file with Microsoft YaHei font for all cells and auto-width headers.
    """
    field_names = list(schema.model_fields.keys())

    # 创建 DataFrame
    if data_list:
        data_dicts = [item.model_dump() for item in data_list]
        user_export_df = pd.DataFrame(data_dicts, columns=field_names)
    else:
        user_export_df = pd.DataFrame(columns=field_names)

    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"{file_name}_{timestamp}.xlsx"
    stream = io.BytesIO()

    try:
        with pd.ExcelWriter(stream, engine="openpyxl") as writer:
            user_export_df.to_excel(writer, index=False, sheet_name=filename)

            # 获取 worksheet
            workbook = writer.book
            worksheet = workbook.active  # 获取默认工作表

            # 设置字体 & 表头加粗
            yahei_font = Font(name="Microsoft YaHei")
            bold_font = Font(name="Microsoft YaHei", bold=True)

            for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
                for cell in row:
                    cell.font = yahei_font if row_idx > 1 else bold_font  # 第一行加粗

            # 计算最佳列宽
            for col_idx, col_name in enumerate(user_export_df.columns, start=1):
                column_letter = get_column_letter(col_idx)
                max_length = max(
                    user_export_df[col_name].astype(str).map(len).max() if not user_export_df.empty else 0,
                    len(col_name)
                )
                worksheet.column_dimensions[column_letter].width = max(max_length + 2, 15)  # 适当增加宽度

        # 保存到流
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        raise
